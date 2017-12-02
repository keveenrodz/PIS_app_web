from django.shortcuts import render
from app.persona.models import Persona,Estudiante,Asesor,Directivo
from app.persona.forms import EstudianteForm,AsesorForm,DirectivoForm
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.core.urlresolvers import reverse_lazy

from django.views.generic.base import TemplateView
from openpyxl import Workbook
from django.http.response import HttpResponse

# Create your views here.

################################################
############## CLASES DE GESTION ESTUDIANTE LISAR, EDITAR, ELIMINAR Y BORRAR ###############
class ReporteEstudiantesExcel(TemplateView):
    model = Estudiante

    def get(self, request, *args, **kwargs,):

        estudiantes = self.model.objects.all()
        print(estudiantes)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'LISTA ESTUDIANTES'
        ws.merge_cells('B1:F1')
        ws['B3'] = 'CEDULA'
        ws['C3'] = 'NOMBRE'
        ws['D3'] = 'PRIMER APELLIDO'
        ws['E3'] = 'SEGUNDO APELLIDO'
        ws['F3'] = 'AREA ESTUDIO'

        cont = 4
        for Estudiante in estudiantes:
            ws.cell(row=cont, column=2).value = Estudiante.cedula
            ws.cell(row=cont, column=3).value = Estudiante.nombre
            ws.cell(row=cont, column=4).value = Estudiante.primer_apellido
            ws.cell(row=cont, column=5).value = Estudiante.segundo_apellido
            ws.cell(row=cont, column=6).value = Estudiante.area.nombre
            cont = cont + 1
        nombre_archivo = "ReporteEstudiantes.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class EstudianteList(ListView):
    model = Estudiante
    template_name = 'persona/estudiante_list.html'

class EstudianteCreate(CreateView):
    model = Estudiante
    form_class = EstudianteForm
    template_name = 'persona/estudiante_form.html'
    success_url = reverse_lazy('persona:estudiante_listar')

class EstudianteUpdate(UpdateView):
    model = Estudiante
    form_class = EstudianteForm
    template_name = 'persona/estudiante_form.html'
    success_url = reverse_lazy('persona:estudiante_listar')

class EstudianteDelete(DeleteView):
    model = Estudiante
    template_name = 'persona/estudiante_delete.html'
    success_url = reverse_lazy('persona:estudiante_listar')

################################################
############## CLASES DE GESTION ASESOR LISAR, EDITAR, ELIMINAR Y BORRAR ###############

class AsesorList(ListView):
    model = Asesor
    template_name = 'persona/asesor_list.html'

class AsesorCreate(CreateView):
    model = Asesor
    form_class = AsesorForm
    template_name = 'persona/asesor_form.html'
    success_url = reverse_lazy('persona:asesor_listar')

class AsesorUpdate(UpdateView):
    model = Asesor
    form_class = AsesorForm
    template_name = 'persona/asesor_form.html'
    success_url = reverse_lazy('persona:asesor_listar')

class AsesorDelete(DeleteView):
    model = Asesor
    template_name = 'persona/asesor_delete.html'
    success_url = reverse_lazy('persona:asesor_listar')

################################################
############## CLASES DE GESTION ASESOR LISAR, EDITAR, ELIMINAR Y BORRAR ###############

class DirectivoList(ListView):
    model = Directivo
    template_name = 'persona/directivo_list.html'

class DirectivoCreate(CreateView):
    model = Directivo
    form_class = DirectivoForm
    template_name = 'persona/directivo_form.html'
    success_url = reverse_lazy('persona:directivo_listar')

class DirectivoUpdate(UpdateView):
    model = Directivo
    form_class = DirectivoForm
    template_name = 'persona/directivo_form.html'
    success_url = reverse_lazy('persona:directivo_listar')

class DirectivoDelete(DeleteView):
    model = Directivo
    template_name = 'persona/directivo_delete.html'
    success_url = reverse_lazy('persona:directivo_listar')

class DirectivoVista(TemplateView):
    template_name = 'persona/directivo_vista.html'

    def get_context_data(self, **kwargs):
        context = super(DirectivoVista, self).get_context_data(**kwargs)
        context['latest_articles'] = Directivo.objects.all()[:5]
        return context
