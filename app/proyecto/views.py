from django.shortcuts import render, redirect
from app.proyecto.forms import ProyectoForm, ProyectoForm2, ProyectoForm3
from app.proyecto.models import ProyectoGrado
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.core.urlresolvers import reverse_lazy
from django.shortcuts import render
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from django.http.response import HttpResponse

# Create your views here.

def post_listar(request):#####Proyectos del directivo (todos)
    current_user = request.user
    queryset = ProyectoGrado.objects.all()
    context = {
        "object_list": queryset,
        "title": "List"
    }
    return render(request, "proyecto/proyecto_list.html", context)

def post_listarasignados(request):####Proyectos del profesor
    current_user = request.user
    if ProyectoGrado.objects.filter(correo_profesor=str(current_user.email)):
        queryset = ProyectoGrado.objects.extra(where=["correo_profesor='" + str(current_user.email) + "'"])
    else:
        queryset = ProyectoGrado.objects.none()
    context = {
        "object_list": queryset,
        "title": "List"
    }
    return render(request, "proyecto/proyecto_list2.html", context)

def post_misproyectos(request):####Proyectos del profesor
    current_user = request.user
    if ProyectoGrado.objects.filter(correo_alumno=str(current_user.email)):
        queryset = ProyectoGrado.objects.extra(where=["correo_alumno='" + str(current_user.email) + "'"])
    else:
        queryset = ProyectoGrado.objects.none()
    context = {
        "object_list": queryset,
        "title": "List"
    }
    return render(request, "proyecto/proyecto_list3.html", context)

class ProyectoCreate(CreateView):
    model = ProyectoGrado
    form_class = ProyectoForm
    template_name = 'proyecto/proyecto_form.html'
    success_url = reverse_lazy('proyecto:proyecto_listar')


class ProyectoUpdate(UpdateView):
    model = ProyectoGrado
    form_class = ProyectoForm
    template_name = 'proyecto/proyecto_form.html'
    success_url = reverse_lazy('proyecto:proyecto_listar')


class ProyectoUpdate2(UpdateView):####Estudiante solo sube o baja archivos
    model = ProyectoGrado
    form_class = ProyectoForm2
    template_name = 'proyecto/proyecto_form.html'
    success_url = reverse_lazy('proyecto:proyecto_listarmisproyectos')

class ProyectoUpdate3(UpdateView):####Profesor o asesor
    model = ProyectoGrado
    form_class = ProyectoForm3
    template_name = 'proyecto/proyecto_form.html'
    success_url = reverse_lazy('proyecto:proyecto_listarasginados')

class ProyectoDelete(DeleteView):
    model = ProyectoGrado
    template_name = 'proyecto/proyecto_delete.html'
    success_url = reverse_lazy('proyecto:proyecto_listar')

class ReporteProyectoExcel(TemplateView):
    model = ProyectoGrado

    def get(self, request, *args, **kwargs):

        proyecto = self.model.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'LISTA PROYECTO'
        ws.merge_cells('B1:K1')
        ws['B3'] = 'NOMBRE'
        ws['C3'] = 'FECHA INSCRIPCION'
        ws['D3'] = 'ALUMNO CEDULA'
        ws['E3'] = 'PROFESOR CEDULA'
        ws['F3'] = 'AREA ESTUDIO'
        ws['G3'] = 'COMENTARIO'
        ws['H3'] = 'NOTA 1'
        ws['I3'] = 'NOTA 2'
        ws['J3'] = 'NOTA 3'
        ws['K3'] = 'NOTA 4'

        cont = 4
        for ProyectoGrado in proyecto:
            ws.cell(row=cont, column=2).value = ProyectoGrado.nombre
            ws.cell(row=cont, column=3).value = ProyectoGrado.fecha_incripcion
            ws.cell(row=cont, column=4).value = ProyectoGrado.alumno.cedula
            ws.cell(row=cont, column=5).value = ProyectoGrado.profesor.cedula
            ws.cell(row=cont, column=6).value = ProyectoGrado.area.nombre
            ws.cell(row=cont, column=7).value = ProyectoGrado.comentario
            ws.cell(row=cont, column=8).value = ProyectoGrado.nota1
            ws.cell(row=cont, column=9).value = ProyectoGrado.nota2
            ws.cell(row=cont, column=10).value = ProyectoGrado.nota3
            ws.cell(row=cont, column=11).value = ProyectoGrado.nota4
            cont = cont + 1
        nombre_archivo = "ReporteProyecto.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response